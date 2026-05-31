"""Stage 2: Translate extracted text rows into Korean with slide context."""

import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import EXCEL_HEADERS, EXTRACTED_DIR, TRANSLATED_DIR
from translator import translate_text


ILLEGAL_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
SAVE_EVERY = 10
MAX_CONTEXT_CHARS = 3000
COL_WIDTHS = [8, 10, 28, 12, 45, 45, 45, 45, 35, 24]
SLIDE_COLORS = ["EBF3FB", "FEF9E7", "E8F8F5", "FDF2F8", "F0F3F4", "FDFEFE", "EAF2F8", "F9EBEA"]


def clean(value):
    """Remove Excel-illegal control characters from a value."""
    return ILLEGAL_CHARS.sub("", value) if isinstance(value, str) else value


def select_file():
    """Prompt the user to select an extracted workbook."""
    files = sorted(f for f in os.listdir(EXTRACTED_DIR) if f.lower().endswith(".xlsx"))
    if not files:
        print(f"[error] No .xlsx files found in {EXTRACTED_DIR}. Run 1_extract.py first.")
        return None

    print("\nSelect extracted workbook to translate:")
    for idx, filename in enumerate(files, 1):
        print(f"  {idx}. {filename}")

    try:
        return files[int(input("Select number: ").strip()) - 1]
    except (ValueError, IndexError):
        print("[error] Invalid selection.")
        return None


def header_index(header, name):
    """Return an optional header index."""
    if name not in header:
        return None
    return header.index(name)


def get_cell(row, header, name, default=""):
    """Return a row value by header name, falling back to default."""
    idx = header_index(header, name)
    if idx is None or idx >= len(row) or row[idx] is None:
        return default
    return row[idx]


def load_rows(path):
    """Load translation rows from an extracted or translated workbook."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        final_translation = get_cell(row, header, "최종 번역")
        draft = get_cell(row, header, "1차 번역 (KO)", final_translation)
        rows.append(
            {
                "slide_num": get_cell(row, header, "슬라이드"),
                "shape_id": get_cell(row, header, "Shape ID"),
                "shape_name": get_cell(row, header, "Shape 이름"),
                "shape_type": get_cell(row, header, "유형"),
                "full_text": get_cell(row, header, "원문 (FR)"),
                "translated": draft or "",
                "claude_review": get_cell(row, header, "Claude 감수"),
                "final_translation": final_translation or draft or "",
                "matched_str": get_cell(row, header, "매칭 용어"),
                "flag": get_cell(row, header, "오류 플래그"),
            }
        )
    return rows


def build_slide_contexts(rows):
    """Group source text by slide number for contextual translation."""
    contexts = {}
    for row in rows:
        text = str(row["full_text"]).strip()
        if text:
            contexts.setdefault(row["slide_num"], []).append(text)
    return contexts


def make_context_for(item, slide_texts):
    """Build same-slide context, excluding the current text item."""
    text = str(item["full_text"]).strip()
    others = [candidate for candidate in slide_texts if candidate != text]
    context = "\n".join(others)
    if len(context) > MAX_CONTEXT_CHARS:
        return context[:MAX_CONTEXT_CHARS] + " ..."
    return context


def style_sheet(ws):
    """Apply the standard translated workbook formatting."""
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_fill = PatternFill("solid", start_color="1F4E79")
    flag_font = Font(name="Arial", size=10, color="C0392B")
    flag_fill = PatternFill("solid", start_color="FADBD8")

    for col, (header, width) in enumerate(zip(EXCEL_HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_idx in range(2, ws.max_row + 1):
        slide_num = ws.cell(row=row_idx, column=1).value or 1
        row_fill = PatternFill("solid", start_color=SLIDE_COLORS[(slide_num - 1) % len(SLIDE_COLORS)])
        for col in range(1, len(EXCEL_HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            if col == 10 and cell.value:
                cell.font = flag_font
                cell.fill = flag_fill
            else:
                cell.font = Font(name="Arial", size=10)
                cell.fill = row_fill
            cell.alignment = left if col in {5, 6, 7, 8, 9} else center
        ws.row_dimensions[row_idx].height = 60

    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    ws.freeze_panes = "A2"


def save_workbook(rows, out_path):
    """Save translated rows to an Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "translated"
    ws.append(EXCEL_HEADERS)

    for row in rows:
        ws.append(
            [
                row["slide_num"],
                row["shape_id"],
                clean(row["shape_name"]),
                row["shape_type"],
                clean(row["full_text"]),
                clean(row["translated"]),
                clean(row["claude_review"]),
                clean(row["final_translation"] or row["translated"]),
                clean(row["matched_str"]),
                clean(row["flag"]),
            ]
        )

    style_sheet(ws)
    wb.save(out_path)


def main():
    """Run the interactive translation stage."""
    extracted_file = select_file()
    if not extracted_file:
        return

    base = extracted_file.replace("_extracted.xlsx", "")
    extracted_path = os.path.join(EXTRACTED_DIR, extracted_file)
    out_path = os.path.join(TRANSLATED_DIR, f"{base}_translated.xlsx")
    rows = load_rows(out_path if os.path.exists(out_path) else extracted_path)

    todo = [idx for idx, row in enumerate(rows) if not str(row.get("translated", "")).strip()]
    print(f"\n[translate] total={len(rows)} done={len(rows) - len(todo)} todo={len(todo)}")
    if not todo:
        print("[done] Nothing to translate.")
        return

    slide_contexts = build_slide_contexts(rows)
    processed = 0
    for idx in todo:
        row = rows[idx]
        matched_terms = str(row["matched_str"]).splitlines() if row["matched_str"] else []
        slide_context = make_context_for(row, slide_contexts.get(row["slide_num"], []))
        print(f"[translate] {idx + 1}/{len(rows)} slide {row['slide_num']} / {row['shape_name']}")
        draft = translate_text(row["full_text"], matched_terms, slide_context)
        row["translated"] = draft
        row["final_translation"] = draft
        processed += 1

        if processed % SAVE_EVERY == 0:
            save_workbook(rows, out_path)
            print(f"[save] progress saved to {out_path}")

    save_workbook(rows, out_path)
    print(f"\n[done] Saved: {out_path}")


if __name__ == "__main__":
    main()
