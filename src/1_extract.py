"""Stage 1: Extract text from PPTX into Excel workbooks."""

import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import EXCEL_HEADERS, EXTRACTED_DIR, INPUT_DIR
from extractor import extract_from_pptx


ILLEGAL_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
COL_WIDTHS = [8, 10, 28, 12, 60, 45, 45, 45, 35, 24]
SLIDE_COLORS = ["EBF3FB", "FEF9E7", "E8F8F5", "FDF2F8", "F0F3F4", "FDFEFE", "EAF2F8", "F9EBEA"]


def clean(value):
    """Remove Excel-illegal control characters from a value."""
    return ILLEGAL_CHARS.sub("", value) if isinstance(value, str) else value


def select_file(folder, extension, prompt):
    """Prompt the user to select a file from a folder."""
    files = sorted(f for f in os.listdir(folder) if f.lower().endswith(extension))
    if not files:
        print(f"[error] No {extension} files found in {folder}")
        return None

    print(f"\n{prompt}")
    for idx, filename in enumerate(files, 1):
        print(f"  {idx}. {filename}")

    try:
        return files[int(input("Select number: ").strip()) - 1]
    except (ValueError, IndexError):
        print("[error] Invalid selection.")
        return None


def style_sheet(ws):
    """Apply the standard extraction workbook formatting."""
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_fill = PatternFill("solid", start_color="1F4E79")

    for col, (header, width) in enumerate(zip(EXCEL_HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    for row in ws.iter_rows(min_row=2):
        row_fill = PatternFill("solid", start_color=SLIDE_COLORS[(row[0].value - 1) % len(SLIDE_COLORS)])
        for col, cell in enumerate(row, 1):
            cell.font = Font(name="Arial", size=10)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = left if col in {5, 6, 7, 8, 9} else center

    ws.row_dimensions[1].height = 30
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 48
    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    ws.freeze_panes = "A2"


def save_extracted(items, pptx_filename):
    """Save extracted items to the 1_extracted workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "extracted"
    ws.append(EXCEL_HEADERS)

    for item in items:
        ws.append(
            [
                item["slide_num"],
                item["shape_id"],
                clean(item["shape_name"]),
                item["shape_type"],
                clean(item["full_text"]),
                "",
                "",
                "",
                clean("\n".join(item["matched_terms"])),
                clean(item["flag"]),
            ]
        )

    style_sheet(ws)
    out_path = os.path.join(EXTRACTED_DIR, f"{os.path.splitext(pptx_filename)[0]}_extracted.xlsx")
    wb.save(out_path)
    return out_path


def print_stats(items):
    """Print extraction summary statistics."""
    textboxes = sum(item["shape_type"] == "textbox" for item in items)
    cells = sum(item["shape_type"] == "table_cell" for item in items)
    matched = sum(bool(item["matched_terms"]) for item in items)
    flagged = sum(bool(item["flag"]) for item in items)
    slides = max((item["slide_num"] for item in items), default=0)
    print(f"\nSlides: {slides}")
    print(f"Textboxes: {textboxes} / table cells: {cells}")
    print(f"Glossary matches: {matched} / flags: {flagged}")
    print(f"Total extracted: {len(items)}")


def main():
    """Run the interactive extraction stage."""
    pptx_filename = select_file(INPUT_DIR, ".pptx", "Select PPTX to extract:")
    if not pptx_filename:
        return

    out_path = os.path.join(EXTRACTED_DIR, f"{os.path.splitext(pptx_filename)[0]}_extracted.xlsx")
    if os.path.exists(out_path):
        answer = input(f"\nExisting extraction found. Overwrite {out_path}? (y/n): ").strip().lower()
        if answer != "y":
            print("Cancelled.")
            return

    print(f"\n[extract] {pptx_filename}")
    items, _ = extract_from_pptx(os.path.join(INPUT_DIR, pptx_filename))
    if not items:
        print("[error] No text found.")
        return

    print_stats(items)
    saved_path = save_extracted(items, pptx_filename)
    print(f"\n[done] Saved: {saved_path}")


if __name__ == "__main__":
    main()
